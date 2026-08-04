"""
Parse the Kopkar kantin July 2026 workbook (LP + PB sheets) per the
kopkar-html-dashboard-report skill rules in SKILL.md §11.
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
import json
import os

XLSX = r"C:\Users\bradl\Documents\Jimmy Greei Ganap\kopkar-report-generator\raw-reports\laporan penjualan kantin bulan juli 2026.xlsx"


def num(v, default=0.0):
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).replace('.', '').replace(',', '.').strip()
        return float(s)
    except ValueError:
        return default


def norm_cat(c):
    if not c:
        return 'LAIN2'
    s = str(c).strip().upper()
    syn = {'LAIN-LAIN': 'LAIN2', 'LAINNYA': 'LAIN2', 'LAIN2.': 'LAIN2'}
    return syn.get(s, s)


wb = openpyxl.load_workbook(XLSX, data_only=True)
print("SHEETS:", wb.sheetnames)

# Match PB / LP by prefix
pb = next(wb[n] for n in wb.sheetnames if n.strip().upper().startswith('PB'))
lp = next(wb[n] for n in wb.sheetnames if n.strip().upper().startswith('LP'))
print("PB sheet:", pb.title)
print("LP sheet:", lp.title)

# ---------- LP ----------
print("\n--- LP header rows 12-14 (cols B..P = 2..16) ---")
for r in (12, 13, 14, 15):
    vals = [lp.cell(row=r, column=c).value for c in range(2, 17)]
    print(r, vals)

# Parse LP from row 14
produk = []
for row in lp.iter_rows(min_row=14, values_only=True):
    # row[0] is column A (None), row[2] is NAMA PRODUK (col C)
    nama = row[2] if len(row) > 2 else None
    if not nama or not str(nama).strip():
        continue
    if str(nama).strip().upper() in ('TOTAL', 'JUMLAH', 'GRAND TOTAL'):
        break
    kategori = norm_cat(row[4]) if len(row) > 4 else 'LAIN2'
    hpp = num(row[5]) if len(row) > 5 else 0.0
    jual = num(row[6]) if len(row) > 6 else 0.0
    qty_awal = num(row[7]) if len(row) > 7 else 0.0
    t_kredit = num(row[8]) if len(row) > 8 else 0.0
    t_tunai = num(row[9]) if len(row) > 9 else 0.0
    sisa = num(row[10]) if len(row) > 10 else 0.0
    rp_kredit = num(row[11]) if len(row) > 11 else 0.0
    rp_tunai = num(row[12]) if len(row) > 12 else 0.0
    modal = num(row[13]) if len(row) > 13 else 0.0
    laba = num(row[14]) if len(row) > 14 else 0.0
    opname = num(row[15]) if len(row) > 15 else 0.0
    produk.append({
        'nama': str(nama).strip(),
        'kategori': kategori,
        'satuan': str(row[3]).strip() if len(row) > 3 and row[3] else '',
        'hpp': hpp, 'harga_jual': jual,
        'qty_awal': qty_awal, 't_kredit': t_kredit, 't_tunai': t_tunai,
        'terjual': t_kredit + t_tunai,
        'sisa': sisa,
        'rp_kredit': rp_kredit, 'rp_tunai': rp_tunai,
        'omzet': rp_kredit + rp_tunai,
        'modal': modal, 'laba': laba, 'opname': opname,
    })

print(f"\nParsed {len(produk)} produk rows")

# Reconciliation check (critical per §11)
omzet_total = sum(p['omzet'] for p in produk)
modal_total = sum(p['modal'] for p in produk)
laba_total = sum(p['laba'] for p in produk)
rp_kredit_total = sum(p['rp_kredit'] for p in produk)
rp_tunai_total = sum(p['rp_tunai'] for p in produk)

print(f"\n--- TOTALS ---")
print(f"Omzet (kredit+tunai) : Rp {omzet_total:,.0f}")
print(f"  Kredit             : Rp {rp_kredit_total:,.0f} ({rp_kredit_total/omzet_total*100:.1f}%)")
print(f"  Tunai              : Rp {rp_tunai_total:,.0f} ({rp_tunai_total/omzet_total*100:.1f}%)")
print(f"Modal                : Rp {modal_total:,.0f}")
print(f"Laba kotor (col O)   : Rp {laba_total:,.0f}")
print(f"Reconciliation (omzet-modal vs laba): diff = {(omzet_total - modal_total) - laba_total:,.0f}")
assert abs((omzet_total - modal_total) - laba_total) < 1, "RECON FAILED"
print("✓ Reconciliation passed")

# Per kategori
print(f"\n--- PER KATEGORI ---")
kateg = {}
for p in produk:
    k = p['kategori']
    kateg.setdefault(k, {'count': 0, 'omzet': 0, 'modal': 0, 'laba': 0, 'terjual': 0, 'rp_kredit': 0, 'rp_tunai': 0})
    kateg[k]['count'] += 1
    kateg[k]['omzet'] += p['omzet']
    kateg[k]['modal'] += p['modal']
    kateg[k]['laba'] += p['laba']
    kateg[k]['terjual'] += p['terjual']
    kateg[k]['rp_kredit'] += p['rp_kredit']
    kateg[k]['rp_tunai'] += p['rp_tunai']

for k in sorted(kateg, key=lambda x: -kateg[x]['omzet']):
    v = kateg[k]
    margin = v['laba'] / v['omzet'] * 100 if v['omzet'] else 0
    print(f"  {k:8s} n={v['count']:3d}  omzet={v['omzet']:13,.0f} ({v['omzet']/omzet_total*100:5.1f}%)  "
          f"modal={v['modal']:12,.0f}  laba={v['laba']:12,.0f}  margin={margin:5.1f}%")

# Dead stock: TERJUAL = 0
dead = [p for p in produk if p['terjual'] == 0]
dead_with_stock = [p for p in dead if p['sisa'] > 0]
print(f"\n--- DEAD STOCK ---")
print(f"Produk dengan TERJUAL = 0: {len(dead)}")
print(f"  di antaranya masih ada stok (sisa>0): {len(dead_with_stock)}")
print(f"  di antaranya tanpa stok (sisa=0): {len(dead) - len(dead_with_stock)}")
dead_stock_value = sum(p['hpp'] * p['sisa'] for p in dead_with_stock)
print(f"Nilai modal terikat di dead-stock berstok: Rp {dead_stock_value:,.0f}")

# Top produk by omzet
top_omzet = sorted([p for p in produk if p['omzet'] > 0], key=lambda x: -x['omzet'])[:10]
print(f"\n--- TOP 10 PRODUK BY OMZET ---")
for i, p in enumerate(top_omzet, 1):
    margin = p['laba'] / p['omzet'] * 100 if p['omzet'] else 0
    print(f"  {i:2d}. {p['nama'][:30]:30s} [{p['kategori']:8s}] omzet={p['omzet']:12,.0f}  laba={p['laba']:10,.0f}  margin={margin:5.1f}%")

# Top produk by laba
top_laba = sorted([p for p in produk if p['laba'] > 0], key=lambda x: -x['laba'])[:10]
print(f"\n--- TOP 10 PRODUK BY LABA ---")
for i, p in enumerate(top_laba, 1):
    margin = p['laba'] / p['omzet'] * 100 if p['omzet'] else 0
    print(f"  {i:2d}. {p['nama'][:30]:30s} [{p['kategori']:8s}] laba={p['laba']:10,.0f}  omzet={p['omzet']:12,.0f}  margin={margin:5.1f}%")

# Stock value
stock_value = sum(p['hpp'] * p['sisa'] for p in produk)
print(f"\n--- STOCK ---")
print(f"Nilai sisa stok (HPP × sisa): Rp {stock_value:,.0f}")
print(f"Setara {stock_value/omzet_total*100:.1f}% omzet sebulan")

# Opname reconciliation
opname_total = sum(p['opname'] for p in produk)
print(f"STOCK OPNAME total (col P): Rp {opname_total:,.0f}")
print(f"Sisa stok value (HPP×sisa) : Rp {stock_value:,.0f}")
print(f"Selisih: Rp {opname_total - stock_value:,.0f}")

# Highest/lowest margin produk (with meaningful sales, omzet>50k)
meaningful = [p for p in produk if p['omzet'] > 50000]
by_margin = sorted(meaningful, key=lambda x: -(x['laba'] / x['omzet'] if x['omzet'] else 0))
print(f"\n--- MARGIN PER PRODUK (omzet > 50k) ---")
print("Top 5 (margin tertinggi):")
for p in by_margin[:5]:
    m = p['laba'] / p['omzet'] * 100 if p['omzet'] else 0
    print(f"  {p['nama'][:30]:30s} margin={m:5.1f}%  jual={p['harga_jual']:,.0f}  hpp={p['hpp']:,.0f}")
print("Bottom 5 (margin terendah):")
for p in by_margin[-5:]:
    m = p['laba'] / p['omzet'] * 100 if p['omzet'] else 0
    print(f"  {p['nama'][:30]:30s} margin={m:5.1f}%  jual={p['harga_jual']:,.0f}  hpp={p['hpp']:,.0f}")

# Items active
active = sum(1 for p in produk if p['terjual'] > 0)
print(f"\n--- ITEM STATUS ---")
print(f"Produk aktif (terjual>0): {active}/{len(produk)}")
print(f"Produk mati  (terjual=0): {len(dead)}/{len(produk)}")

# Save aggregated data as JSON for the generator script
agg = {
    'produk': produk,
    'totals': {
        'omzet': omzet_total,
        'modal': modal_total,
        'laba': laba_total,
        'rp_kredit': rp_kredit_total,
        'rp_tunai': rp_tunai_total,
        'margin_pct': laba_total / omzet_total * 100 if omzet_total else 0,
        'hpp_ratio_pct': modal_total / omzet_total * 100 if omzet_total else 0,
        'stock_value': stock_value,
        'opname_value': opname_total,
        'n_produk': len(produk),
        'n_active': active,
        'n_dead': len(dead),
        'n_dead_with_stock': len(dead_with_stock),
        'n_dead_no_stock': len(dead) - len(dead_with_stock),
        'dead_stock_value': dead_stock_value,
        'kredit_share_pct': rp_kredit_total / omzet_total * 100 if omzet_total else 0,
    },
    'kategori': {k: v for k, v in kateg.items()},
}
out = os.path.join(os.path.dirname(__file__), '_juli_2026.json')
with open(out, 'w', encoding='utf-8') as f:
    json.dump(agg, f, ensure_ascii=False, indent=2)
print(f"\nSaved aggregated → {out}")
